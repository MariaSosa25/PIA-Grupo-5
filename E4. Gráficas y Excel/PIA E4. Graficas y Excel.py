# Semana 4 PIA
# Grupo 5

import requests
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np

# Esta función permite validar, mediante expresiones regulares, que el usuario ingrese únicamente digitos en el ID
# Evita errores en las solicitudes a la API
def validar_id(asteroid_id):
    return re.fullmatch(r"\d+", asteroid_id) is not None

class Manejo_de_api:

    # Función que conecta el programa con el API de NEO de la NASA
    # Determina si el asteroide es peligroso o no
    def determinar_peligro(API_KEY, asteroid_id):
        url= f"https://api.nasa.gov/neo/rest/v1/neo/{asteroid_id}?api_key={API_KEY}"
        respuesta= requests.get(url)
        if respuesta.status_code == 200:
            datos = respuesta.json()
            nombre = datos["name"]
            print(f"\nNombre del asteroide {asteroid_id}: {nombre}")
            es_peligroso = datos["is_potentially_hazardous_asteroid"]
            if es_peligroso==True:
                print(f"\n{nombre} es un asteroide potencialmente peligroso.")
            else:
                print(f"\n{nombre} NO es un asteroide potencialmente peligroso.")
        else:
            print(f"Error {respuesta.status_code}: no se pudo conectar con la API. Verifica que el ID sea válido.")

    # Esta función permite que el usuario visualice datos de un asteroide
    # Le da la opción de crear un archivo txt con los datos
    def consultar_datos(API_KEY, asteroid_id):
        url= f"https://api.nasa.gov/neo/rest/v1/neo/{asteroid_id}?api_key={API_KEY}"
        respuesta= requests.get(url)
        if respuesta.status_code == 200:
            datos = respuesta.json()
            # Evaluación de limpieza de datos:
            # Los datos obtenidos del JSON están bien estructurados y no presentan campos ruidosos o inconsistencias.
            # Por lo tanto, no se requiere limpieza adicional.

            # Se crea un diccionario 
            datos_detallados = dict()
            datos_detallados["NOMBRE"]= datos["name"]
            datos_detallados["ID"]= datos["id"]
            datos_detallados["MAGNITUD_ABSOLUTA_H"]= datos["absolute_magnitude_h"]
            datos_detallados["DIÁMETRO_ESTIMADO"]= datos["estimated_diameter"]["kilometers"] # Datos compuestos de otro dict

            nombre = datos.get("name_limited", datos["name"])

            # Se imprime el diccionario
            for x,y in datos_detallados.items():
                print(f"\n{x} --> {y}")
                print("__________")
            while True:
                opcion= int(input(f"\nDesea guardar la información de {nombre} en un archivo de texto? [1-Si/2-No]: "))
                if opcion==1:
                    with open (f"Datos_{asteroid_id}.txt","w") as file:
                        for x,y in datos_detallados.items():
                            file.write(f"{x} --> {y}.\n")
                    print("Listo!")
                    break
                elif opcion==2:
                    print("Regresando al menú principal…")
                    break
                else:
                    print("Opcion invalida. ¡Pruebe de nuevo!")
        else:
            print(f"Error {respuesta.status_code}: no se pudo conectar con la API. Verifica que el ID sea válido.")

    # Esta función permite al usuario buscar la cantidad de IDs que desee según un intervalo propuesto.
    # Regresa una lista de IDs
    def lista_ids(API_KEY, inicial, final):
        while True:
            if inicial < 0 or inicial > final or final > 39312:
                print("\nRango inválido. Intenta con un intervalo válido dentro de 0 y 39312")
                try:
                    inicial = int(input("Introducir valor inicial del intervalo: "))
                    final = int(input("Introducir valor final del intervalo: "))
                except ValueError:
                    print("Por favor ingrese solo números.")
            else:
                break
        asteroides_totales = []
        page = 0
        while len(asteroides_totales) <= final:
            url = f"https://api.nasa.gov/neo/rest/v1/neo/browse?page={page}&size=100&api_key={API_KEY}"
            respuesta = requests.get(url)
            if respuesta.status_code == 200:
                datos = respuesta.json()
                nuevos = datos["near_earth_objects"]
                # Evaluación de limpieza de datos:
                # La API ya proporciona una lista clara de objetos con campos estructurados.
                # No se identificaron errores, valores nulos o datos redundantes, por lo que no fue necesaria una limpieza adicional.

                asteroides_totales.extend(nuevos)
                
                total_pages = datos["page"]["total_pages"]
                current_page = datos["page"]["number"]

                if current_page >= total_pages - 1:
                    break

                # Se navega mediante las páginas del API para obtener los IDs solicitados por el usuario
                page += 1
            else:
                print(f"Error {respuesta.status_code}: no se pudo conectar con la API. Verifica que el ID sea válido.")
                return

        # Se imprime la lista de los IDs encontrados    
        print("\nIDs disponibles en el intervalo solicitado:")
        for i in range(inicial, final + 1):
            asteroide = asteroides_totales[i]
            if (inicial == 0):
                print(f"{i+1}. ID: {asteroide['id']} | Nombre: {asteroide['name']}")
            else:
                print(f"{i}. ID: {asteroide['id']} | Nombre: {asteroide['name']}")

        while True:
                opcion= int(input(f"\nDesea guardar la lista de IDs en un archivo de texto? [1-Si/2-No]: "))
                if opcion==1:
                    with open (f"Asteroides_IDs.txt","w") as file:
                        for i in range(inicial, final + 1):
                            asteroide = asteroides_totales[i]
                            if (inicial == 0):
                                file.write(f"\n{i+1}. ID: {asteroide['id']} | Nombre: {asteroide['name']}")
                            else:
                                file.write(f"\n{i}. ID: {asteroide['id']} | Nombre: {asteroide['name']}")
                    print("Listo!")
                    break
                elif opcion==2:
                    print("Regresando al menú principal…")
                    break
                else:
                    print("Opcion invalida. ¡Pruebe de nuevo!")

# Se creo otra clase para trabajar solo con el excel y las graficas
class Manejo_excel:
    # Se conecta a una api general
    def conectar_api_general():
        API_KEY= "0AMkNAnXRuPBPKGLJWGrqDFQfcb4qbniMnvQVc0x"
        url=f"https://api.nasa.gov/neo/rest/v1/neo/browse?api_key={API_KEY}"
        respuesta= requests.get(url)
        if respuesta.status_code == 200:
            # Crea un excel de los datos de la api
            info_dict = respuesta.json()
            libro = Workbook()
            pagina = libro.active
            pagina.title = "datos_generales"
            # Los encabezados 
            pagina.append(["ID", "NOMBRE", "MAGNITUD H", "PRIMER FECHA DE OBSERVACION", "JTI", "CLASE DE ORBITA"])
            for info in info_dict["near_earth_objects"]: 
                hilera = [info["id"], info["name_limited"], info["absolute_magnitude_h"], info["orbital_data"]["first_observation_date"],
                          info["orbital_data"]["jupiter_tisserand_invariant"],info["orbital_data"]["orbit_class"]["orbit_class_type"]]
                
                pagina.append(hilera) 
            libro.save("Informacion_general.xlsx")
            print("")
        else:
            print(f"Error {respuesta.status_code}: no se pudo conectar con la API. Verifica que el ID sea válido.")

    def visualizar_magnitudes():
        # Verifica que el usuario tenga el excel
        while True:
            try:
                libro = load_workbook("Informacion_general.xlsx")
            except FileNotFoundError:
                Manejo_excel.conectar_api_general()
            finally:
                break
        
        hoja= libro.active
        nombres= list()
        magnitudes= list()
        rango= hoja["B2":"C21"]
        # Guarda la info. para utilizarla en las graficas
        for celda in rango:
            for objeto in celda:
                if objeto.column==2: # Numero de columna
                    nombres.append(objeto.value) 
                else:
                    valor = str(objeto.value)
                    # Se valida la magnitud
                    if re.fullmatch(r"\d+\.\d+", valor):
                        magnitudes.append(float(valor))
        # Organiza los datos para la utilizacion de la libreria plt
        x= np.array(nombres)
        y= np.array(magnitudes)
        colores= ["lightsteelblue", "powderblue", "cadetblue"]
        plt.bar(x,y,color= colores)
        plt.xticks(fontsize=5)
        plt.xlabel("NEOs (nombres)") 
        plt.ylabel("Magnitudes H") 
        plt.title("Gráfica de comparación de magnitud absoluta H")
        plt.show()

    def clase_orbita():
        while True:
            try:
                libro = load_workbook("Informacion_general.xlsx")
            except FileNotFoundError:
                Manejo_excel.conectar_api_general()
            finally:
                break
        hoja=libro.active
        co= list() # co --> clase orbita
        rango=hoja["F2":"F21"]
        for celda in rango:
            for objeto in celda:
                if objeto.column==6:
                    co.append(objeto.value)
        suma_apo=0
        suma_amo=0
        # Contabiliza los NEOs APO y AMO
        for elem in co:
            # Se usan expresiones regulares para validar la clase
            if re.fullmatch(r"^(AMO|APO)$", str(elem)):
                if elem=="APO":
                    suma_apo+= 1
                else:
                    suma_amo+= 1
        y= np.array([suma_apo, suma_amo])
        sig= ["APO (cruzan la órbita de la Tierra)", "AMO  (rozan la órbita terrestre pero no la cruzan)"]
        colores =["lightsalmon","peachpuff"]
        plt.pie(y, labels= sig, colors= colores)
        plt.title("Gráfica de abundancia de NEOs AMO y APO en la base de datos")
        plt.show()

    def fecha():
        while True:
            try:
                libro = load_workbook("Informacion_general.xlsx")
            except FileNotFoundError:
                Manejo_excel.conectar_api_general()
            finally:
                break
        
        hoja= libro.active
        nombres= list()
        fechas= list()
        rango= hoja["B2":"D21"]

        for celda in rango:
            for objeto in celda:
                if objeto.column==2: # Numero de columna
                    nombres.append(objeto.value) 
                elif objeto.column==4:
                    # Como el formato de la fecha es aaaa-mm-dd
                    # Las expresiones regulares agrupan solo el año
                    patron = r"(\d{4})"
                    resultado = re.search(patron, str(objeto.value))
                    if resultado:
                        fechas.append(resultado.group(0))
        x= np.array(nombres)
        y= np.array(fechas)
        plt.scatter(x,y, marker='*', color="slateblue")
        plt.xticks(fontsize=8)
        plt.xlabel("NEOs (nombres)") 
        plt.ylabel("Año") 
        plt.title("Gráfica de año de primer observación")
        plt.show()


    def TJ():
        while True:
            try:
                libro = load_workbook("Informacion_general.xlsx")
            except FileNotFoundError:
                Manejo_excel.conectar_api_general()
            finally:
                break
        
        hoja= libro.active
        nombres= list()
        PTJ= list()
        rango= hoja["B2":"E21"]

        for celda in rango:
            for objeto in celda:
                if objeto.column==2: # Numero de columna
                    nombres.append(objeto.value) 
                elif objeto.column==5:
                    valor = str(objeto.value)  # Convierte cualquier tipo a string
                    if re.fullmatch(r"\d+\.\d+", valor):  # Valida que sea un decimal como '4.512'
                        PTJ.append(float(valor))  # Convierte y agrega
        x= np.array(nombres)
        y= np.array(PTJ)
        colores = ["palevioletred", "lavenderblush", "thistle"]
        plt.barh(x,y, color= colores)
        plt.xticks(fontsize=5)
        plt.xlabel("Invariante de Tisserand de Júpiter")
        plt.ylabel("NEOs (nombres)") 
        plt.title("Gráfica parámetros de Tisserand relacionados con Júpiter")
        plt.show()        

# Programa inicial
API_KEY= "0AMkNAnXRuPBPKGLJWGrqDFQfcb4qbniMnvQVc0x"
print("""¡Bienvenido!
Este programa tiene como objetivo proporcionar datos relevantes sobre objetos NEO (Near Earth Objects).
Elija una opción de nuestro menú para comenzar.
""")
menu = """
-------------------------------------Menú-------------------------------------
| 1. Determinar si un asteroide es potencialmente peligroso (se necesita ID) |
| 2. Consultar sobre NEOs (se necesita ID)                                   |
| 3. Consultar IDs en la base de datos                                       |
| 4. Análisis gráfico de la base de datos                                    |
| 5. Salir                                                                   |
------------------------------------------------------------------------------
"""
while True:
    print(menu)
    # Aquí se valida que el usuario no ingrese caracteres que no sean enteros
    try:
        op = int(input("Ingrese opción deseada: "))
    except ValueError:
        print("La opción debe de ser un número. ¡Pruebe de nuevo!")
    else:
        if op==1:
            # Ejemplos de id: 2001566, 2001221, 2001580…
            asteroid_id= input("Ingresar ID: ")
            # Se valida el ID con expresiones regulares
            if validar_id(asteroid_id):
                Manejo_de_api.determinar_peligro(API_KEY, asteroid_id)
            else:
                print("ID inválido. Debe contener solo dígitos.")
            input("\nPresiona la tecla enter para continuar. :D")
        elif op==2:
            asteroid_id= input("Ingresar ID: ")
            # Se valida el ID con expresiones regulares
            if validar_id(asteroid_id):
                Manejo_de_api.consultar_datos(API_KEY, asteroid_id)
            else:
                print("ID inválido. Debe contener solo dígitos.")
            input("\nPresiona la tecla enter para continuar. :D")
        elif op==3:
            print("El sistema de datos cuenta con 39312 elementos, por esta razón le recomendamos visualizar pocas IDs a la vez!")
            print("Tenga en cuenta que, entre mayores sean los números, más se tardará en procesar la información.")
            # Se validan los intervalos haciendo uso de excepciones
            try:
                inicial = int(input("Introducir valor inicial del intervalo: "))
                final = int(input("Introducir valor final del intervalo: "))
            except ValueError:
                print("Por favor ingrese solo números.")
                input("\nPresiona la tecla enter para continuar. :D")
            else:
                print("Buscando IDs…")
                Manejo_de_api.lista_ids(API_KEY, inicial, final)
                input("\nPresiona la tecla enter para continuar. :D")
        elif op==4:
            menu2="""
---------------------Menú de gráficas-------------------------
| 1. Comparación de magnitudes H (gráfica de barras)         |
| 2. Clase de órbita (gráfica de pastel)                     |
| 3. Primer fecha de observación (gráfica de dispersión)     |
| 4. Parámetros de Tisserand relacionados con Júpiter        |
|    (gráfica de barras horizontales)                        |
| 5. Volver al menú principal                                |
--------------------------------------------------------------
"""
            while True:
                print(menu2)
                op2= int(input("Ingrese el gráfico que quiera visualizar [1-5]: "))
                if op2==1:
                    Manejo_excel.conectar_api_general()
                    Manejo_excel.visualizar_magnitudes()
                elif op2==2:
                    Manejo_excel.conectar_api_general()
                    Manejo_excel.clase_orbita()
                elif op2==3:
                    Manejo_excel.conectar_api_general()
                    Manejo_excel.fecha()
                elif op2==4:
                    Manejo_excel.conectar_api_general()
                    Manejo_excel.TJ()
                elif op2==5:
                    print("Regresando...")
                    break
                else:
                    print("Opcion invalida. ¡Pruebe de nuevo!")
                
        elif op==5:
            print("Salida del usuario.\nHASTA LUEGO!")
            break
        else:
            print("Opcion invalida. ¡Pruebe de nuevo!")
